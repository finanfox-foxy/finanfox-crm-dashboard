#!/usr/bin/env python3
"""Extract 2025 BBVA transactions, categorize, and merge into financial_data.json"""

import openpyxl
import json
import re
from collections import defaultdict

# ── Read Excel ──
wb = openpyxl.load_workbook('/Users/foxy/Desktop/Backup Telegram/files/Movimientos BBVA 2025.xlsx', data_only=True)
ws = wb['Historico']

transactions = []  # list of {date, amount, concepto, beneficiario, observaciones}

for row_idx in range(17, ws.max_row + 1):
    date_str = ws.cell(row=row_idx, column=3).value
    concepto = ws.cell(row=row_idx, column=6).value
    beneficiario = ws.cell(row=row_idx, column=7).value
    observaciones = ws.cell(row=row_idx, column=8).value
    importe_str = ws.cell(row=row_idx, column=9).value

    if not date_str or not importe_str:
        continue

    try:
        parts = str(date_str).split('/')
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        date_obj = f"{year:04d}-{month:02d}"
    except:
        continue

    try:
        amount = float(str(importe_str).replace(',', '').replace(' ', ''))
    except:
        continue

    concepto_str = str(concepto).lower() if concepto else ''
    beneficiario_str = str(beneficiario).lower() if beneficiario else ''
    observaciones_str = str(observaciones).lower() if observaciones else ''
    combined = concepto_str + ' ' + beneficiario_str + ' ' + observaciones_str

    transactions.append({
        'month': date_obj,
        'amount': amount,
        'concepto': str(concepto or ''),
        'beneficiario': str(beneficiario or ''),
        'observaciones': str(observaciones or ''),
        'combined': combined
    })

print(f"Total transactions extracted: {len(transactions)}")

# ── Categorize Income ──
def categorize_income(combined):
    c = combined
    # Precious Metals / PMC
    if 'pmc' in c or 'precious metals' in c or 'intermediaci' in c:
        return 'PMC (Intermediación)'
    # Monefit with factura
    if 'monefit' in c and 'factura' in c:
        return 'Monefit'
    # Creditstar = Real Money (matching existing convention)
    if 'creditstar' in c:
        return 'Real Money'
    # Silver & Gold with liquidacion/liq/ajustes
    if 'silver' in c and 'gold' in c and ('liquidacion' in c or ' liq' in c or 'ajustes' in c):
        return 'Silver / Gold'
    # SG Patrimonio / SGPM
    if 'sg patrimonio' in c or 'sgpm' in c:
        return 'Silver / Gold'
    # Santa Lucia (insurance)
    if 'santa lucia' in c or 'santalucia' in c:
        return 'Seguros y Cías'
    if 'asesores de seguros' in c:
        return 'Seguros y Cías'
    # Sunday Marketplace
    if 'sunday marketplace' in c or 'sunday mp' in c or 'sundaymarketplace' in c:
        return 'Sunday Marketplace'
    # Stripe
    if 'stripe' in c:
        return 'Stripe'
    # IronBridge
    if 'ironbridge' in c or 'iron bridge' in c:
        return 'IronBridge'
    # Ahorro y Proteccion / Mapfre
    if 'ahorro y proteccion' in c:
        return 'Seguros y Cías'
    if 'mapfre' in c:
        return 'Seguros y Cías'
    # Real Money other sources
    if 'real money' in c or 'rcm' in c:
        return 'Real Money'
    if 'mortgage direct' in c:
        return 'Otros Ingresos'
    return None  # not clearly income → will be 'Otros Ingresos'

# ── Categorize Expense ──
def categorize_expense(combined, amount_abs, beneficiario_lower, concepto_lower, observaciones_lower):
    c = combined

    # Nóminas and similar
    if 'nomina' in c or 'nómina' in c or 'n0mina' in c:
        return 'Nóminas'
    if 'alberto prieto' in c:
        return 'Nóminas'
    if 'cristina cabello' in c:
        return 'Nóminas'
    if 'alejandro caballero' in c:
        return 'Nóminas'
    if 'jaime becerra' in c or 'jaime israel' in c:
        return 'Nóminas'
    if 'jose orrequia' in c or 'jose javier' in c:
        return 'Nóminas'

    # Seguridad Social
    if 'seguridad social' in c or 'tgss' in c:
        return 'Seguridad Social'

    # Impuestos
    if 'hacienda' in c or 'aeat' in c or 'impuesto' in c:
        return 'Impuestos'

    # Transfers/investments to SG/Silver+Gold accounts → Gastos financieros
    # These are outbound transfers to investment accounts (CCU, etc.)
    if 'silver' in c and 'gold' in c and ('ccu' in c or 'transfer' in c):
        return 'Gastos financieros'
    if 'sg patrimonio' in c and ('ccu' in c or 'reserva' in c or 'transfer' in c):
        return 'Gastos financieros'

    # Monefit Card/Smartsaver transfers
    if 'monefit' in c and 'card' in c:
        return 'Gastos financieros'
    if 'monefit' in c and ('smart' in c or 'ahorro' in c or 'saving' in c):
        return 'Gastos financieros'

    # Marketing
    if 'google' in c and ('ads' in c or 'cloud' in c):
        return 'Marketing'
    if 'marketinhouse' in c:
        return 'Marketing'
    if 'digital product services' in c or 'foxfield' in c:
        return 'Marketing'
    if 'referral-factory' in c or 'tremendous' in c:
        return 'Marketing'
    if 'flake agency' in c:
        return 'Marketing'
    if 'hubspot' in c:
        return 'Marketing'
    if 'atres advertising' in c or 'atresmedia' in c:
        return 'Marketing'
    if 'contentcreator' in c or 'content' in c:
        return 'Marketing'
    if 'ediciones horo' in c:
        return 'Marketing'
    if 'neointec' in c:
        return 'Proveedores'
    if 'empresa nacional de innovacion' in c or 'enisa' in c:
        return 'Marketing'

    # Suscripciones
    if 'microsoft' in c:
        return 'Suscripciones'
    if 'xerintel' in c:
        return 'Suscripciones'
    if 'wealth reader' in c:
        return 'Suscripciones'
    if 'qubiq' in c:
        return 'Suscripciones'
    if 'soluciones web' in c:
        return 'Suscripciones'
    if 'o2 fibra' in c or 'telefonica' in c:
        return 'Suscripciones'
    if 'boluda' in c or 'zadarma' in c:
        return 'Suscripciones'
    if 'railway' in c:
        return 'Suscripciones'
    if 'qloudea' in c:
        return 'Suscripciones'
    if 'odoo' in c:
        return 'Suscripciones'
    if 'zoho' in c:
        return 'Suscripciones'

    # Seguros
    if 'mapfre' in c:
        return 'Seguros'
    if 'aegon' in c:
        return 'Seguros'
    if 'caser' in c:
        return 'Seguros'

    # Proveedores
    if 'home for sale' in c:
        return 'Proveedores'
    if 'mabor consult' in c:
        return 'Proveedores'

    # Gastos varios specific
    if 'asesoria' in c or 'luis m' in c:
        return 'Gastos varios'
    if 'american express' in c:
        return 'Gastos varios'
    if 'fundacion bahia' in c:
        return 'Gastos varios'
    if 'kc sales' in c:
        return 'Gastos varios'
    if 'growfix' in c:
        return 'Gastos varios'
    if 'northgate' in c:
        return 'Gastos varios'
    if 'hotel' in c:
        return 'Gastos varios'

    return 'Gastos varios'


# ── Aggregate by month ──
months_data = defaultdict(lambda: {
    'income': defaultdict(float),
    'expenses': defaultdict(float),
    'income_total': 0.0,
    'expenses_total': 0.0,
    'count': 0
})

for t in transactions:
    month = t['month']
    amount = t['amount']
    combined = t['combined']
    beneficiario_lower = t['beneficiario'].lower()
    concepto_lower = t['concepto'].lower()
    observaciones_lower = t['observaciones'].lower()

    months_data[month]['count'] += 1

    if amount > 0:
        cat = categorize_income(combined)
        if not cat:
            cat = 'Otros Ingresos'
        months_data[month]['income'][cat] += amount
        months_data[month]['income_total'] += amount
    else:
        abs_amount = abs(amount)
        cat = categorize_expense(combined, abs_amount, beneficiario_lower, concepto_lower, observaciones_lower)
        months_data[month]['expenses'][cat] += abs_amount
        months_data[month]['expenses_total'] += abs_amount

# Round all values to 2 decimal places
for m in months_data:
    for k in ['income', 'expenses']:
        months_data[m][k] = {cat: round(val, 2) for cat, val in months_data[m][k].items()}
    months_data[m]['income_total'] = round(months_data[m]['income_total'], 2)
    months_data[m]['expenses_total'] = round(months_data[m]['expenses_total'], 2)

# Print summary
total_2025_income = 0
total_2025_expenses = 0
for m in sorted(months_data.keys()):
    d = months_data[m]
    margin = d['income_total'] - d['expenses_total']
    total_2025_income += d['income_total']
    total_2025_expenses += d['expenses_total']
    print(f"{m}: income={d['income_total']:>10.2f} expenses={d['expenses_total']:>10.2f} margin={margin:>10.2f} txns={d['count']}")
    print(f"   Income: {dict(sorted(d['income'].items(), key=lambda x: -x[1]))}")
    print(f"   Expenses: {dict(sorted(d['expenses'].items(), key=lambda x: -x[1]))}")

print(f"\n2025 TOTAL: income={total_2025_income:.2f} expenses={total_2025_expenses:.2f} margin={total_2025_income - total_2025_expenses:.2f}")

# ── Build 2025 months ──
def build_month_entry(month_data):
    income_cats = month_data['income']
    expense_cats = month_data['expenses']
    income_total = month_data['income_total']
    expenses_total = month_data['expenses_total']
    margin = round(income_total - expenses_total, 2)

    if income_cats:
        big_inc_name = max(income_cats, key=income_cats.get)
        big_inc_amt = income_cats[big_inc_name]
        big_inc_pct = round(big_inc_amt / income_total * 100, 1) if income_total > 0 else 0
    else:
        big_inc_name, big_inc_amt, big_inc_pct = None, 0, 0

    if expense_cats:
        big_exp_name = max(expense_cats, key=expense_cats.get)
        big_exp_amt = expense_cats[big_exp_name]
        big_exp_pct = round(big_exp_amt / expenses_total * 100, 1) if expenses_total > 0 else 0
    else:
        big_exp_name, big_exp_amt, big_exp_pct = None, 0, 0

    return {
        'income': {
            'total': income_total,
            'categories': dict(sorted(income_cats.items(), key=lambda x: -x[1]))
        },
        'expenses': {
            'total': expenses_total,
            'categories': dict(sorted(expense_cats.items(), key=lambda x: -x[1]))
        },
        'margin': margin,
        'biggest_income_category': {
            'name': big_inc_name,
            'amount': round(big_inc_amt, 2),
            'pct': big_inc_pct
        },
        'biggest_expense_category': {
            'name': big_exp_name,
            'amount': round(big_exp_amt, 2),
            'pct': big_exp_pct
        },
        'transaction_count': month_data['count']
    }

new_2025_months = {}
for m in sorted(months_data.keys()):
    new_2025_months[m] = build_month_entry(months_data[m])

# ── Load existing JSON ──
with open('/Users/foxy/.openclaw/workspace/finanfox-crm-dashboard/data/financial_data.json') as f:
    existing = json.load(f)

# Keep existing 2026 months
existing_2026 = {k: v for k, v in existing['months'].items() if k.startswith('2026')}

# Merge
merged_months = {}
merged_months.update(new_2025_months)
merged_months.update(existing_2026)

# ── Recalculate ──
all_months_sorted = sorted(merged_months.keys())

# monthly_evolution
monthly_evolution = []
for m in all_months_sorted:
    d = merged_months[m]
    monthly_evolution.append({
        'month': m,
        'income': d['income']['total'],
        'expenses': d['expenses']['total'],
        'margin': d['margin']
    })

# ytd and mtd
ytd_income = sum(merged_months[m]['income']['total'] for m in all_months_sorted)
ytd_expenses = sum(merged_months[m]['expenses']['total'] for m in all_months_sorted)
last_month = all_months_sorted[-1]
mtd_data = merged_months[last_month]

# category_summary
income_cats_overall = defaultdict(lambda: {'total': 0.0, 'months': {}})
expense_cats_overall = defaultdict(lambda: {'total': 0.0, 'months': {}})

for m in all_months_sorted:
    d = merged_months[m]
    for cat, val in d['income']['categories'].items():
        income_cats_overall[cat]['total'] += val
        income_cats_overall[cat]['months'][m] = val
    for cat, val in d['expenses']['categories'].items():
        expense_cats_overall[cat]['total'] += val
        expense_cats_overall[cat]['months'][m] = val

# Fill missing months with 0
for cat in income_cats_overall:
    months_have = income_cats_overall[cat]['months']
    for m in all_months_sorted:
        if m not in months_have:
            months_have[m] = 0.0

for cat in expense_cats_overall:
    months_have = expense_cats_overall[cat]['months']
    for m in all_months_sorted:
        if m not in months_have:
            months_have[m] = 0.0

# Round and sort everything
for cat in income_cats_overall:
    d = income_cats_overall[cat]
    d['total'] = round(d['total'], 2)
    d['months'] = dict(sorted(d['months'].items()))
    d['pct'] = round(d['total'] / ytd_income * 100, 1) if ytd_income > 0 else 0

for cat in expense_cats_overall:
    d = expense_cats_overall[cat]
    d['total'] = round(d['total'], 2)
    d['months'] = dict(sorted(d['months'].items()))
    d['pct'] = round(d['total'] / ytd_expenses * 100, 1) if ytd_expenses > 0 else 0

income_cats_overall = dict(sorted(income_cats_overall.items(), key=lambda x: -x[1]['total']))
expense_cats_overall = dict(sorted(expense_cats_overall.items(), key=lambda x: -x[1]['total']))

# ── Build output ──
from datetime import datetime

output = {
    'timestamp': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
    'generated_at': datetime.utcnow().strftime('%Y-%m-%d'),
    'period_start': '2025-01-01',
    'period_end': '2026-05-18',
    'account_name': existing['account_name'],
    'iban': existing['iban'],
    'months': merged_months,
    'ytd': {
        'income_total': round(ytd_income, 2),
        'expenses_total': round(ytd_expenses, 2),
        'margin': round(ytd_income - ytd_expenses, 2)
    },
    'mtd': {
        'month': last_month,
        'income_total': mtd_data['income']['total'],
        'expenses_total': mtd_data['expenses']['total'],
        'margin': mtd_data['margin']
    },
    'category_summary': {
        'income': income_cats_overall,
        'expenses': expense_cats_overall
    },
    'monthly_evolution': monthly_evolution
}

# ── Write ──
with open('/Users/foxy/.openclaw/workspace/finanfox-crm-dashboard/data/financial_data.json', 'w') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"\n✅ Written financial_data.json successfully!")
print(f"Period: {output['period_start']} to {output['period_end']}")
print(f"Total months: {len(all_months_sorted)}")
print(f"2025 months: {[m for m in all_months_sorted if m.startswith('2025')]}")
print(f"2026 months: {[m for m in all_months_sorted if m.startswith('2026')]}")
print(f"YTD Income: {output['ytd']['income_total']:.2f}")
print(f"YTD Expenses: {output['ytd']['expenses_total']:.2f}")
print(f"YTD Margin: {output['ytd']['margin']:.2f}")

# Verify JSON validity
json.dumps(output)  # will throw if invalid
print("✅ JSON valid!")